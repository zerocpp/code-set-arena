"""Teacher-side roster and Stage 1 package import helpers."""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from .constants import KIND_PROBLEMS, ROLE_STUDENT, STAGE1
from .course_validation import validate_stage1_problem_package
from .package_names import PackageNameError, assert_student_archive
from .packages import PackageError, read_package
from .teacher_version_gate import assert_student_package_version_allowed


@dataclass(frozen=True)
class Stage1ImportResult:
    student_number: str
    archive_name: str
    problem_count: int
    version_tag: str = ""
    repaired_runs: int = 0


def load_student_roster_xlsx(path: Path) -> dict[str, dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only in broken environments.
        raise RuntimeError("读取学生名单.xlsx 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    students: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_number = _cell_text(row[0] if len(row) > 0 else "")
        if not student_number:
            continue
        students[student_number] = {
            "student_number": student_number,
            "name": _cell_text(row[1] if len(row) > 1 else ""),
            "class_id": _cell_text(row[2] if len(row) > 2 else ""),
        }
    if not students:
        raise ValueError("学生名单为空")
    return students


def import_stage1_archive(
    root: Path,
    state: dict[str, Any],
    archive_path: Path,
    *,
    reject_duplicate: bool = True,
) -> Stage1ImportResult:
    result, student, problems = validate_stage1_archive(root, state, archive_path)
    submissions = state.setdefault("submissions", {})
    if reject_duplicate and (
        result.student_number in submissions
        or (root / "stage1-submissions/uploads" / archive_path.name).exists()
    ):
        raise ValueError("同名或同学号 Stage 1 包已导入，请先删除原包")

    upload_target = root / "stage1-submissions/uploads" / archive_path.name
    upload_target.parent.mkdir(parents=True, exist_ok=True)
    if archive_path.resolve() != upload_target.resolve():
        shutil.copy2(archive_path, upload_target)
    import_dir = root / "stage1-submissions/imports" / _archive_import_dir_name(archive_path)
    read_package(upload_target, import_dir)
    submissions[result.student_number] = {
        "student": student,
        "problems": problems,
        "archive": archive_path.name,
        "version_tag": result.version_tag,
        "received_at": datetime.now(UTC).isoformat(),
    }
    _set_stage1_status(
        state,
        result.student_number,
        imported=True,
        ok=True,
        archive=archive_path.name,
        version_tag=result.version_tag,
        detail=f"校验通过，题目数 {result.problem_count}",
    )
    return result


def validate_stage1_archive(
    root: Path,
    state: dict[str, Any],
    archive_path: Path,
) -> tuple[Stage1ImportResult, dict[str, Any], list[dict[str, Any]]]:
    manifest, payload = read_package(
        archive_path, root / "stage1-submissions/validation-reports" / _archive_import_dir_name(archive_path)
    )
    _assert_manifest(manifest, ROLE_STUDENT, STAGE1, KIND_PROBLEMS)
    version_tag = str(manifest.get("version_tag") or "").strip()
    assert_student_package_version_allowed(manifest, state.get("settings", {}))
    student = payload.get("student", {})
    student_number = str(student.get("student_number", "")).strip()
    try:
        assert_student_archive(archive_path, student_number, STAGE1, KIND_PROBLEMS)
    except PackageNameError as exc:
        filename_student = _student_number_from_archive_name(archive_path.name)
        raise ValueError(
            f"包名学号 {filename_student} 与包内学号 {student_number} 不一致"
        ) from exc
    roster = state.get("students", {})
    if roster and student_number not in roster:
        raise ValueError(f"{student_number} 未在学生名单中")
    problems = payload.get("problems", [])
    if not isinstance(problems, list):
        raise ValueError("payload problems must be a list")
    repaired_before = _count_runs_with_final_text(problems)
    validate_stage1_problem_package(problems)
    repaired_after = _count_runs_with_final_text(problems)
    return (
        Stage1ImportResult(
            student_number=student_number,
            archive_name=archive_path.name,
            problem_count=len(problems),
            version_tag=version_tag,
            repaired_runs=max(0, repaired_after - repaired_before),
        ),
        student,
        problems,
    )


def revalidate_stage1_archives(root: Path, state: dict[str, Any]) -> dict[str, str]:
    results: dict[str, str] = {}
    uploads = sorted((root / "stage1-submissions/uploads").glob("*.tar.gz"))
    state["submissions"] = {}
    for archive in uploads:
        try:
            result = import_stage1_archive(root, state, archive, reject_duplicate=False)
        except Exception as exc:  # noqa: BLE001 - status table needs the exact validation error.
            student_number = _student_number_from_archive_name(archive.name)
            _set_stage1_status(
                state,
                student_number,
                imported=True,
                ok=False,
                archive=archive.name,
                version_tag=_stage1_archive_version_tag(root, archive),
                detail=str(exc),
            )
            results[archive.name] = str(exc)
        else:
            results[archive.name] = "ok"
            _set_stage1_status(
                state,
                result.student_number,
                imported=True,
                ok=True,
                archive=archive.name,
                version_tag=result.version_tag,
                detail=f"校验通过，题目数 {result.problem_count}",
            )
    return results


def stage1_roster_rows(state: dict[str, Any]) -> list[dict[str, Any]]:
    students = state.get("students", {})
    submissions = state.get("submissions", {})
    statuses = state.get("stage1_package_status", {})
    student_numbers = sorted(set(students) | set(submissions) | set(statuses))
    rows = []
    for student_number in student_numbers:
        student = students.get(student_number) or submissions.get(student_number, {}).get("student", {})
        status = statuses.get(student_number, {})
        submission = submissions.get(student_number)
        archive = (
            submission.get("archive", status.get("archive", ""))
            if submission
            else status.get("archive", "")
        )
        rows.append(
            {
                "student_number": student_number,
                "name": student.get("name", ""),
                "class_id": student.get("class_id", ""),
                "imported": bool(submission) or bool(status.get("imported")),
                "archive": archive,
                "version_tag": (
                    submission.get("version_tag", status.get("version_tag", ""))
                    if submission
                    else status.get("version_tag", "")
                ),
                "validation_ok": bool(status.get("ok")) if status else bool(submission),
                "validation_detail": status.get("detail", "未导入"),
                "has_submission": bool(submission),
                "can_delete": bool(archive),
            }
        )
    return rows


def stage1_invalid_statuses(state: dict[str, Any]) -> dict[str, str]:
    invalid: dict[str, str] = {}
    for student_number, status in state.get("stage1_package_status", {}).items():
        if status.get("archive") and not status.get("ok"):
            invalid[student_number] = str(status.get("detail", "校验失败"))
    return invalid


def missing_roster_students(state: dict[str, Any]) -> list[dict[str, str]]:
    students = state.get("students", {})
    submissions = state.get("submissions", {})
    return [students[number] for number in sorted(set(students) - set(submissions))]


def _set_stage1_status(
    state: dict[str, Any],
    student_number: str,
    *,
    imported: bool,
    ok: bool,
    archive: str,
    detail: str,
    version_tag: str = "",
) -> None:
    state.setdefault("stage1_package_status", {})[student_number] = {
        "imported": imported,
        "ok": ok,
        "archive": archive,
        "version_tag": version_tag,
        "detail": detail,
        "checked_at": datetime.now(UTC).isoformat(),
    }


def _assert_manifest(manifest: dict[str, Any], role: str, stage: str, kind: str) -> None:
    if manifest.get("package_role") != role:
        raise PackageError("manifest package_role 不匹配")
    if manifest.get("package_stage") != stage:
        raise PackageError("manifest package_stage 不匹配")
    if manifest.get("package_kind") != kind:
        raise PackageError("manifest package_kind 不匹配")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _archive_import_dir_name(path: Path) -> str:
    name = path.name
    return name[: -len(".tar.gz")] if name.endswith(".tar.gz") else path.stem


def _stage1_archive_version_tag(root: Path, archive_path: Path) -> str:
    try:
        manifest, _ = read_package(
            archive_path,
            root / "stage1-submissions/validation-reports" / _archive_import_dir_name(archive_path),
        )
    except Exception:  # noqa: BLE001 - best effort for the status table.
        return ""
    return str(manifest.get("version_tag") or "").strip()


def _student_number_from_archive_name(name: str) -> str:
    return name.split("-student-", 1)[0] if "-student-" in name else name


def _count_runs_with_final_text(problems: list[dict[str, Any]]) -> int:
    return sum(
        1
        for problem in problems
        for run in problem.get("run_records", [])
        if run.get("raw_response") or run.get("extracted_code")
    )
